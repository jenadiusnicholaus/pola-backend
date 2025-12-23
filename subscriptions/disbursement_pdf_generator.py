"""
Disbursement PDF Generator

Generates professional PDF receipts for disbursements that can be downloaded by admins.
"""

import base64
from io import BytesIO
from decimal import Decimal
from datetime import datetime
from xhtml2pdf import pisa
from django.template.loader import render_to_string
from django.utils import timezone


class DisbursementPDFGenerator:
    """Generate PDF receipts for disbursements"""
    
    @staticmethod
    def generate_pdf(disbursement):
        """
        Generate a PDF receipt for a disbursement
        
        Args:
            disbursement: Disbursement model instance
            
        Returns:
            dict: {
                'pdf_base64': str,  # Base64 encoded PDF
                'filename': str,    # Suggested filename
                'size_bytes': int   # File size in bytes
            }
        """
        # Prepare data for template
        context = {
            'disbursement': disbursement,
            'recipient': disbursement.recipient,
            'generated_at': timezone.now(),
            'consultant_earnings': disbursement.consultant_earnings.all(),
            'uploader_earnings': disbursement.uploader_earnings.all(),
            'total_consultant_earnings': sum(
                e.net_earnings for e in disbursement.consultant_earnings.all()
            ),
            'total_uploader_earnings': sum(
                e.net_earnings for e in disbursement.uploader_earnings.all()
            ),
        }
        
        # Render HTML template
        html_content = DisbursementPDFGenerator._render_html(context)
        
        # Convert HTML to PDF
        pdf_buffer = BytesIO()
        pisa_status = pisa.CreatePDF(
            html_content,
            dest=pdf_buffer,
            encoding='utf-8'
        )
        
        if pisa_status.err:
            raise Exception(f"PDF generation failed with error code: {pisa_status.err}")
        
        # Get PDF bytes
        pdf_bytes = pdf_buffer.getvalue()
        pdf_buffer.close()
        
        # Encode to base64
        pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
        
        # Generate filename
        filename = f"disbursement_{disbursement.external_reference}_{datetime.now().strftime('%Y%m%d')}.pdf"
        
        return {
            'pdf_base64': pdf_base64,
            'filename': filename,
            'size_bytes': len(pdf_bytes),
            'mimetype': 'application/pdf'
        }
    
    @staticmethod
    def _render_html(context):
        """Render HTML template for PDF"""
        disbursement = context['disbursement']
        recipient = context['recipient']
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{
                    size: A4;
                    margin: 2cm;
                }}
                body {{
                    font-family: Arial, sans-serif;
                    font-size: 11pt;
                    color: #333;
                    line-height: 1.6;
                }}
                .header {{
                    text-align: center;
                    margin-bottom: 30px;
                    border-bottom: 3px solid #2563eb;
                    padding-bottom: 20px;
                }}
                .header h1 {{
                    color: #2563eb;
                    margin: 0;
                    font-size: 24pt;
                }}
                .header p {{
                    color: #666;
                    margin: 5px 0;
                }}
                .info-box {{
                    background-color: #f3f4f6;
                    padding: 15px;
                    border-radius: 5px;
                    margin: 20px 0;
                }}
                .info-row {{
                    display: flex;
                    justify-content: space-between;
                    margin: 8px 0;
                }}
                .info-label {{
                    font-weight: bold;
                    color: #666;
                    width: 40%;
                }}
                .info-value {{
                    width: 60%;
                    text-align: right;
                }}
                .amount-box {{
                    background-color: #dcfce7;
                    border: 2px solid #16a34a;
                    padding: 20px;
                    text-align: center;
                    margin: 20px 0;
                    border-radius: 5px;
                }}
                .amount-box h2 {{
                    color: #16a34a;
                    margin: 0;
                    font-size: 28pt;
                }}
                .amount-label {{
                    color: #666;
                    font-size: 12pt;
                    margin: 5px 0;
                }}
                table {{
                    width: 100%;
                    border-collapse: collapse;
                    margin: 20px 0;
                }}
                th {{
                    background-color: #2563eb;
                    color: white;
                    padding: 10px;
                    text-align: left;
                    font-weight: bold;
                }}
                td {{
                    padding: 8px 10px;
                    border-bottom: 1px solid #e5e7eb;
                }}
                tr:nth-child(even) {{
                    background-color: #f9fafb;
                }}
                .totals-row {{
                    font-weight: bold;
                    background-color: #e5e7eb !important;
                }}
                .footer {{
                    margin-top: 40px;
                    padding-top: 20px;
                    border-top: 2px solid #e5e7eb;
                    font-size: 9pt;
                    color: #666;
                    text-align: center;
                }}
                .status-badge {{
                    display: inline-block;
                    padding: 5px 15px;
                    border-radius: 20px;
                    font-weight: bold;
                    font-size: 10pt;
                }}
                .status-pending {{
                    background-color: #fef3c7;
                    color: #92400e;
                }}
                .status-processing {{
                    background-color: #dbeafe;
                    color: #1e40af;
                }}
                .status-completed {{
                    background-color: #dcfce7;
                    color: #166534;
                }}
                .status-failed {{
                    background-color: #fee2e2;
                    color: #991b1b;
                }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>POLA LEGAL SERVICES</h1>
                <p>Disbursement Receipt</p>
                <p style="font-size: 9pt;">Generated on {context['generated_at'].strftime('%B %d, %Y at %H:%M:%S')}</p>
            </div>
            
            <div class="info-box">
                <div class="info-row">
                    <span class="info-label">Reference Number:</span>
                    <span class="info-value"><strong>{disbursement.external_reference}</strong></span>
                </div>
                <div class="info-row">
                    <span class="info-label">Disbursement Date:</span>
                    <span class="info-value">{disbursement.initiated_at.strftime('%B %d, %Y')}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">Recipient:</span>
                    <span class="info-value">{recipient.get_full_name()}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">Email:</span>
                    <span class="info-value">{recipient.email}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">Phone Number:</span>
                    <span class="info-value">{disbursement.recipient_phone}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">Payment Method:</span>
                    <span class="info-value">{disbursement.get_payment_method_display()}</span>
                </div>
                <div class="info-row">
                    <span class="info-label">Status:</span>
                    <span class="info-value">
                        <span class="status-badge status-{disbursement.status}">
                            {disbursement.get_status_display().upper()}
                        </span>
                    </span>
                </div>
            </div>
            
            <div class="amount-box">
                <p class="amount-label">Total Disbursement Amount</p>
                <h2>TZS {disbursement.amount:,.2f}</h2>
            </div>
        """
        
        # Add consultant earnings table if any
        consultant_earnings = context['consultant_earnings']
        if consultant_earnings:
            html += """
            <h3 style="color: #2563eb; margin-top: 30px;">Consultation Earnings</h3>
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Service Type</th>
                        <th>Gross Amount</th>
                        <th>Commission</th>
                        <th>Net Earnings</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for earning in consultant_earnings:
                html += f"""
                    <tr>
                        <td>{earning.created_at.strftime('%Y-%m-%d')}</td>
                        <td>{earning.service_type.replace('_', ' ').title()}</td>
                        <td>TZS {earning.gross_amount:,.2f}</td>
                        <td>TZS {earning.platform_commission:,.2f}</td>
                        <td><strong>TZS {earning.net_earnings:,.2f}</strong></td>
                    </tr>
                """
            
            html += f"""
                    <tr class="totals-row">
                        <td colspan="4" style="text-align: right;">SUBTOTAL:</td>
                        <td><strong>TZS {context['total_consultant_earnings']:,.2f}</strong></td>
                    </tr>
                </tbody>
            </table>
            """
        
        # Add uploader earnings table if any
        uploader_earnings = context['uploader_earnings']
        if uploader_earnings:
            html += """
            <h3 style="color: #2563eb; margin-top: 30px;">Content Upload Earnings</h3>
            <table>
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Service Type</th>
                        <th>Gross Amount</th>
                        <th>Commission</th>
                        <th>Net Earnings</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for earning in uploader_earnings:
                material_name = earning.material.title if earning.material else 'Document Template'
                html += f"""
                    <tr>
                        <td>{earning.created_at.strftime('%Y-%m-%d')}</td>
                        <td>{earning.service_type.replace('_', ' ').title()}</td>
                        <td>TZS {earning.gross_amount:,.2f}</td>
                        <td>TZS {earning.platform_commission:,.2f}</td>
                        <td><strong>TZS {earning.net_earnings:,.2f}</strong></td>
                    </tr>
                """
            
            html += f"""
                    <tr class="totals-row">
                        <td colspan="4" style="text-align: right;">SUBTOTAL:</td>
                        <td><strong>TZS {context['total_uploader_earnings']:,.2f}</strong></td>
                    </tr>
                </tbody>
            </table>
            """
        
        # Add notes if any
        if disbursement.notes:
            html += f"""
            <div class="info-box" style="margin-top: 30px;">
                <h4 style="margin-top: 0; color: #666;">Notes:</h4>
                <p style="margin: 0;">{disbursement.notes}</p>
            </div>
            """
        
        # Footer
        html += f"""
            <div class="footer">
                <p><strong>POLA Legal Services Platform</strong></p>
                <p>This is an automatically generated receipt. For inquiries, contact support@pola.com</p>
                <p style="font-size: 8pt; margin-top: 10px;">
                    Document ID: {disbursement.external_reference} | 
                    Generated: {context['generated_at'].strftime('%Y-%m-%d %H:%M:%S')}
                </p>
            </div>
        </body>
        </html>
        """
        
        return html
    
    @staticmethod
    def generate_excel(disbursement):
        """
        Generate an Excel spreadsheet for a disbursement
        
        Args:
            disbursement: Disbursement model instance
            
        Returns:
            dict: {
                'excel_base64': str,  # Base64 encoded Excel file
                'filename': str,      # Suggested filename
                'size_bytes': int     # File size in bytes
            }
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Disbursement"
        
        # Title
        ws.merge_cells('A1:E1')
        title_cell = ws['A1']
        title_cell.value = "POLA LEGAL SERVICES - DISBURSEMENT RECEIPT"
        title_cell.font = Font(size=16, bold=True, color="1F4788")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Basic Info
        row = 3
        info_data = [
            ("Reference Number:", disbursement.external_reference),
            ("Disbursement Date:", disbursement.initiated_at.strftime('%Y-%m-%d %H:%M')),
            ("Recipient:", disbursement.recipient.get_full_name()),
            ("Email:", disbursement.recipient.email),
            ("Phone:", disbursement.recipient_phone),
            ("Payment Method:", disbursement.get_payment_method_display()),
            ("Status:", disbursement.get_status_display()),
            ("Total Amount:", f"TZS {disbursement.amount:,.2f}"),
        ]
        
        for label, value in info_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        row += 2
        
        # Consultant Earnings
        consultant_earnings = disbursement.consultant_earnings.all()
        if consultant_earnings:
            ws[f'A{row}'] = "CONSULTATION EARNINGS"
            ws[f'A{row}'].font = Font(size=12, bold=True, color="1F4788")
            row += 1
            
            # Headers
            headers = ['Date', 'Service Type', 'Gross Amount', 'Commission', 'Net Earnings']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Data
            total = Decimal('0.00')
            for earning in consultant_earnings:
                ws[f'A{row}'] = earning.created_at.strftime('%Y-%m-%d')
                ws[f'B{row}'] = earning.service_type.replace('_', ' ').title()
                ws[f'C{row}'] = float(earning.gross_amount)
                ws[f'D{row}'] = float(earning.platform_commission)
                ws[f'E{row}'] = float(earning.net_earnings)
                total += earning.net_earnings
                row += 1
            
            # Subtotal
            ws[f'D{row}'] = "SUBTOTAL:"
            ws[f'D{row}'].font = Font(bold=True)
            ws[f'E{row}'] = float(total)
            ws[f'E{row}'].font = Font(bold=True)
            row += 2
        
        # Uploader Earnings
        uploader_earnings = disbursement.uploader_earnings.all()
        if uploader_earnings:
            ws[f'A{row}'] = "CONTENT UPLOAD EARNINGS"
            ws[f'A{row}'].font = Font(size=12, bold=True, color="1F4788")
            row += 1
            
            # Headers
            headers = ['Date', 'Service Type', 'Gross Amount', 'Commission', 'Net Earnings']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Data
            total = Decimal('0.00')
            for earning in uploader_earnings:
                ws[f'A{row}'] = earning.created_at.strftime('%Y-%m-%d')
                ws[f'B{row}'] = earning.service_type.replace('_', ' ').title()
                ws[f'C{row}'] = float(earning.gross_amount)
                ws[f'D{row}'] = float(earning.platform_commission)
                ws[f'E{row}'] = float(earning.net_earnings)
                total += earning.net_earnings
                row += 1
            
            # Subtotal
            ws[f'D{row}'] = "SUBTOTAL:"
            ws[f'D{row}'].font = Font(bold=True)
            ws[f'E{row}'] = float(total)
            ws[f'E{row}'].font = Font(bold=True)
            row += 2
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()
        excel_buffer.close()
        
        # Encode to base64
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        
        # Generate filename
        filename = f"disbursement_{disbursement.external_reference}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return {
            'excel_base64': excel_base64,
            'filename': filename,
            'size_bytes': len(excel_bytes),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
    
    @staticmethod
    def generate_bulk_excel(disbursements_queryset, title="Disbursements Report"):
        """
        Generate an Excel spreadsheet for multiple disbursements
        
        Args:
            disbursements_queryset: QuerySet of Disbursement objects
            title: Report title
            
        Returns:
            dict: {
                'excel_base64': str,
                'filename': str,
                'size_bytes': int
            }
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from decimal import Decimal
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Disbursements"
        
        # Title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True, color="1F4788")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Generation info
        ws.merge_cells('A2:I2')
        info_cell = ws['A2']
        info_cell.value = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        info_cell.alignment = Alignment(horizontal='center')
        info_cell.font = Font(size=9, italic=True)
        
        # Headers
        row = 4
        headers = [
            'Reference', 'Date', 'Recipient', 'Email', 'Phone', 
            'Amount (TZS)', 'Status', 'Payment Method', 'Type'
        ]
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        row += 1
        
        # Data rows
        total_amount = Decimal('0.00')
        status_counts = {}
        
        for disbursement in disbursements_queryset:
            ws[f'A{row}'] = disbursement.external_reference
            ws[f'B{row}'] = disbursement.initiated_at.strftime('%Y-%m-%d %H:%M')
            ws[f'C{row}'] = disbursement.recipient.get_full_name()
            ws[f'D{row}'] = disbursement.recipient.email
            ws[f'E{row}'] = disbursement.recipient_phone
            ws[f'F{row}'] = float(disbursement.amount)
            ws[f'G{row}'] = disbursement.get_status_display()
            ws[f'H{row}'] = disbursement.get_payment_method_display()
            ws[f'I{row}'] = disbursement.get_disbursement_type_display()
            
            # Apply borders to all cells
            for col in range(1, 10):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
            
            # Color code by status
            status_colors = {
                'pending': 'FEF3C7',
                'processing': 'DBEAFE',
                'completed': 'DCFCE7',
                'failed': 'FEE2E2',
                'cancelled': 'F3F4F6'
            }
            if disbursement.status in status_colors:
                ws[f'G{row}'].fill = PatternFill(
                    start_color=status_colors[disbursement.status],
                    end_color=status_colors[disbursement.status],
                    fill_type="solid"
                )
            
            # Track totals and counts
            total_amount += disbursement.amount
            status_counts[disbursement.status] = status_counts.get(disbursement.status, 0) + 1
            
            row += 1
        
        # Summary section
        row += 1
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12, color="1F4788")
        row += 1
        
        ws[f'A{row}'] = "Total Disbursements:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = disbursements_queryset.count()
        row += 1
        
        ws[f'A{row}'] = "Total Amount (TZS):"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = float(total_amount)
        ws[f'B{row}'].font = Font(bold=True)
        ws[f'B{row}'].number_format = '#,##0.00'
        row += 2
        
        # Status breakdown
        ws[f'A{row}'] = "Status Breakdown:"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        for status_key, count in status_counts.items():
            ws[f'A{row}'] = status_key.title()
            ws[f'B{row}'] = count
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 22  # Reference
        ws.column_dimensions['B'].width = 18  # Date
        ws.column_dimensions['C'].width = 25  # Recipient
        ws.column_dimensions['D'].width = 30  # Email
        ws.column_dimensions['E'].width = 15  # Phone
        ws.column_dimensions['F'].width = 15  # Amount
        ws.column_dimensions['G'].width = 12  # Status
        ws.column_dimensions['H'].width = 15  # Payment Method
        ws.column_dimensions['I'].width = 15  # Type
        
        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_bytes = excel_buffer.getvalue()
        excel_buffer.close()
        
        # Encode to base64
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        
        # Generate filename
        filename = f"disbursements_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return {
            'excel_base64': excel_base64,
            'filename': filename,
            'size_bytes': len(excel_bytes),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
